"""
__author__ = "Amilton Lobo Mendes Júnior"
__email__ = "amilton.mendes@gmail.com"
"""
import os
from openpyxl import load_workbook
from unidecode import unidecode
import pandas as pd
from mysql.connector.errors import IntegrityError


from utils import  prepara_arquivo_download,Conexao,descompacta_arquivo

class CagedDBConfig():
    def __init__(self,usuario='projeto_bd',senha='teste123',base='fundamentos_bd'):
        self.conexao=Conexao(usuario,senha,base)
        self.script_estrangeiras = []


    def cria_tabela(self,con, sheet, sheet_data):
        """
        Tenta remover do banco de dados a tabela com as informações de domínio constantes do arquivo de layout do
        novo CAGED e, em seguida, as recria e carrega os dados

        Parameters
        ----------
        :param con: Conexão com o banco de dados utilizada para a realização das operações no banco
        :param sheet: String com o nome da planilha no arquivo de layout. Esse nome será utilizado como nome da tabela final no BD
        :param sheet_data: Pandas dataframe com as informações da planilha do excel contendo os valores de domínio para esta informação
        :return:
        """
        con = self.conexao.get_con()

        cursor = con.cursor()
        sql = ''
        nome_tabela = unidecode(sheet)
        print('Criando tabela ' + nome_tabela.upper())
        cursor.execute('DROP TABLE IF EXISTS ' + nome_tabela.upper())

        sql += 'CREATE TABLE ' + nome_tabela.upper() + ' (\n'
        for coluna in sheet_data.columns:
            if unidecode(coluna).lower() =='codigo':
                if nome_tabela in ['secao','cbo2002ocupação','subclasse']:
                    sql += '  ' + unidecode(coluna).lower() + ' VARCHAR(7),\n'
                else:
                    sql += '  ' + unidecode(coluna).lower() + ' INTEGER(3),\n'

            else:
                sql += '  ' + unidecode(coluna).lower() + ' VARCHAR(150),\n'
        sql += '  CONSTRAINT pk_' + nome_tabela + ' PRIMARY KEY(' + unidecode(sheet_data.columns[0]).lower() + ')\n'
        sql += ');'
        cursor.execute(sql)
        con.commit()
        cursor.close()

        ##Insere as linhas das planilhas com informações do domínio no banco ( tabelas acessórias )
        sheet_data.rename(columns=lambda s: unidecode(s).lower(),inplace=True)
        engine = self.conexao.get_engine()
        sheet_data.to_sql(nome_tabela,con=engine,if_exists='append',index=False)



    def cria_tabela_principal(self,con, sheet_data):
        """
        Tenta remover do banco de dados a tabela com as informações das operações do Novo CAGED para, em seguida, as recriar

        Parameters
        ----------
        :param con: Conexão com o banco de dados utilizada para a realização das operações no banco
        :param sheet_data: Pandas dataframe com as informações do layout dos dados do novo CAGED
        :return:
        """
        con = self.conexao.get_con()

        cursor = con.cursor()
        sql = ''
        nome_tabela = 'CAGED'
        print('Criando tabela ' + nome_tabela)
        cursor.execute('DROP TABLE IF EXISTS ' + nome_tabela)

        sql += 'CREATE TABLE ' + nome_tabela + ' (\n'
        sql += '    id INT PRIMARY KEY AUTO_INCREMENT \n'
        for coluna in sheet_data['Variável']:
            if unidecode(coluna).lower() in ['secao','cbo2002ocupação','subclasse']:
                sql += ',  ' + unidecode(coluna).lower() + ' VARCHAR(7)\n'
            else:
                sql += ',  ' + unidecode(coluna).lower() + ' INTEGER(3)\n'

            #sql += ',  ' + unidecode(coluna).lower() + ' VARCHAR(50)\n'

        for coluna in sheet_data['Variável']:
            if coluna == 'fonte':
                self.script_estrangeiras.append('ALTER TABLE CAGED ADD FOREIGN KEY (fonte) REFERENCES  FONTE_DESL(codigo)')
            elif (coluna not in ['competência', 'saldomovimentação', 'idade', 'horascontratuais', 'salário']):
                self.script_estrangeiras.append('ALTER TABLE CAGED ADD FOREIGN KEY (' + unidecode(coluna).lower() + ') REFERENCES  ' + unidecode(
                    coluna).upper() + '(codigo)')
#                sql += ',  FOREIGN KEY (' + unidecode(coluna).lower() + ') REFERENCES ' + unidecode(
#                    coluna).upper() + '(codigo) \n'
        sql += ');'
        cursor.execute(sql)
        con.commit()
        cursor.close()

    def create_tables(self,path_layout):
        """
        Cria uma tabela para cada aba da planilha de layout e, para as planilhas diferentes da 'Layout, insere os dados
        :param path_layout:
        :return:
        """
        layout = load_workbook(path_layout)
        con = self.conexao.get_con()

        cursor = con.cursor()
        sql = ''

        nome_tabela = 'CAGED'
        cursor.execute('DROP TABLE IF EXISTS ' + nome_tabela)
        con.commit()
        cursor.close()

        for sheet in layout.sheetnames:

            if sheet != 'Layout':
                df_sheet = pd.read_excel(path_layout, sheet_name=sheet, engine='openpyxl')
                self.cria_tabela(con, sheet, df_sheet)
        df_sheet = pd.read_excel(path_layout, sheet_name='Layout', engine='openpyxl', skiprows=1)
        self.cria_tabela_principal(con, df_sheet)
        con.disconnect()


    def prepara_bases(self):
        '''if prepara_arquivo_download(os.getcwd()+'\config'\
                ,'ftp://ftp.mtps.gov.br/pdet/microdados/NOVO CAGED/Movimentações/Layout Novo Caged Movimentação.xlsx'\
                ,'\layout_caged.xlsx'):
            self.create_tables(os.getcwd()+'\config\layout_caged.xlsx')'''
        self.create_tables(os.getcwd() + '\config\layout_caged.xlsx')
    def insere_dados_caged(self,lista_meses,head=None,uf=0,regiao=0):
        url = 'ftp://ftp.mtps.gov.br/pdet/microdados/NOVO CAGED/Movimentações/2020/Dezembro/CAGEDMOV2020'
        for mes in lista_meses:
            url_mes = url+mes+'.7z'
            #Descomentar
            #prepara_arquivo_download(os.getcwd()+'\\tmp' , url_mes , '\caged_2020'+mes+'7z')
            #descompacta_arquivo(os.getcwd()+'/tmp/caged_2020'+mes+'7z',os.getcwd()+'/data')

            if regiao!= 0:
                df_caged = pd.read_csv(os.getcwd() + '/data/CAGEDMOV2020' + mes + '.txt', sep=';')
                df_caged = df_caged[df_caged['região']==regiao]
                if head != None:
                    df_caged = df_caged[:head]
            elif uf!= 0:
                df_caged = pd.read_csv(os.getcwd() + '/data/CAGEDMOV2020' + mes + '.txt', sep=';')
                df_caged = df_caged[df_caged['uf']==uf]
                if head != None:
                    df_caged = df_caged[:head]

            elif head == None:
                df_caged = pd.read_csv(os.getcwd()+'/data/CAGEDMOV2020'+mes+'.txt',sep=';')
            else:
                df_caged = pd.read_csv(os.getcwd()+'/data/CAGEDMOV2020'+mes+'.txt',sep=';',nrows=head)
            df_caged.rename(columns=lambda s: unidecode(s).lower(), inplace=True)

            # Eliminar subclasse incorreta
            df_caged = df_caged[df_caged['subclasse'] != '8630505']
            engine = self.conexao.get_engine()
            #Código para inserir linhas de 1000 em 1000, para evitar timeout do banco de dados
            print('Arquivo sendo carregado referente ao mês '+str(mes))
            start=0
            #if mes=='01':
            #    start = 335000
            for start_row in range(start, df_caged.shape[0], 1000):
                end_row = min(start_row + 1000, df_caged.shape[0])
                df_parcial = df_caged.iloc[start_row:end_row, :]

                try:
                    df_parcial.to_sql('caged', con=engine, if_exists='append', index=False,chunksize=1)
                except:
                    for i in range(start_row,end_row,1):
                        df_parcial_int = df_caged.iloc[[i]]
                        try:
                            df_parcial_int.to_sql('caged', con=engine, if_exists='append', index=False,chunksize=1)
                        except Exception as ie:
                            print('Erro no registro '+str(i)+' do arquivo de mes '+mes)
                            print('Erro '+str(ie))
        con = self.conexao.get_con()

        cursor = con.cursor()
        for sql in self.script_estrangeiras:
            cursor.execute(sql)
        con.commit()
        cursor.close()

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    caged = CagedDBConfig(base='fundamentos_2')
    caged.prepara_bases()
    caged.insere_dados_caged(['01','02','03','04','05','06','07','08','09','10','11','12'],regiao=1)
